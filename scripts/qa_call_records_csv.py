#!/usr/bin/env python3
"""Export QA call-records to CSV with a classified outcome reason.

Logs into the rustpbx console, queries the call-records listing for the given
phone number(s) and date range, fetches each record's CDR metadata, classifies
the outcome (answered / rang-but-not-answered / busy / exact server error / ...)
and writes a CSV in the same column layout as the console listing plus a
`reason` column.

Usage:
    python3 scripts/qa_call_records_csv.py \
        --number +918071539218 --number +918071539217 \
        --date-from 2026-06-24 --date-to 2026-06-24 \
        --out call_records_today.csv

    # single day shortcut (sets both from/to):
    python3 scripts/qa_call_records_csv.py -n +918071539218 --date 2026-06-24

Auth/host come from flags or env (QA_BASE_URL, QA_IDENTIFIER, QA_PASSWORD).
Nothing here is committed with credentials; defaults match the QA runbook.
"""
import argparse
import csv
import http.cookiejar
import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter
from datetime import datetime, timedelta, timezone

# Display timezone for human-facing timestamps. Stored CDRs are UTC; the console
# default display tz is Asia/Kolkata (IST, UTC+5:30, no DST) so we render to match.
IST = timezone(timedelta(hours=5, minutes=30))

DEFAULT_BASE = os.environ.get("QA_BASE_URL", "http://13.202.41.67:8080")
DEFAULT_USER = os.environ.get("QA_IDENTIFIER", "admin")
DEFAULT_PASS = os.environ.get("QA_PASSWORD", "Admin__123")

CSV_COLUMNS = [
    "started_at", "direction", "from", "to",
    "status", "status_code",
    "ring_secs", "talk_secs", "total_secs",
    "error_reason",
]


def _parse_ts(value):
    """Parse an RFC3339 timestamp with optional nanosecond precision."""
    if not value:
        return None
    trimmed = re.sub(r"(\.\d{6})\d*", r"\1", value).replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(trimmed)
    except ValueError:
        return None


class Client:
    """Tiny session-cookie HTTP client for the console API."""

    def __init__(self, base, identifier, password):
        self.base = base.rstrip("/")
        self.identifier = identifier
        self.password = password
        jar = http.cookiejar.CookieJar()
        # don't auto-follow the 303 on login so we can confirm it
        self.opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(jar)
        )

    def login(self):
        body = (
            f"identifier={urllib.parse.quote(self.identifier)}"
            f"&password={urllib.parse.quote(self.password)}"
        ).encode()
        req = urllib.request.Request(
            f"{self.base}/console/login",
            data=body,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
        )
        try:
            self.opener.open(req)
        except urllib.error.HTTPError as exc:
            # 303 redirect == success; urllib follows it, other codes raise
            if exc.code not in (301, 302, 303):
                raise
        return self

    def post_json(self, path, payload):
        req = urllib.request.Request(
            f"{self.base}{path}",
            data=json.dumps(payload).encode(),
            headers={"Content-Type": "application/json"},
        )
        with self.opener.open(req) as resp:
            return json.loads(resp.read())

    def get_json(self, path):
        with self.opener.open(f"{self.base}{path}") as resp:
            return json.loads(resp.read())


def fetch_records(client, number, date_from, date_to, per_page=500):
    """Page through the listing for one search term."""
    items, page = [], 1
    while True:
        payload = {
            "page": page,
            "per_page": per_page,
            "filters": {
                "q": number,
                "dateFrom": date_from,
                "dateTo": date_to,
            },
        }
        data = client.post_json("/console/call-records", payload)
        items.extend(data.get("items", []))
        if page >= (data.get("total_pages") or 1):
            break
        page += 1
    return items


def classify(meta, rang):
    """Map a CDR metadata blob to a human outcome string."""
    code = meta.get("statusCode")
    status = meta.get("status")
    last_error = meta.get("lastError") or {}
    reason = last_error.get("reason")

    if status == "completed" and code == 200:
        return ""  # answered/normal call has no error reason
    if code == 480:
        return "Rang but not answered" if rang else "Temporarily unavailable (no ring)"
    if code == 487:
        return "Caller cancelled (after ring)" if rang else "Caller cancelled"
    if code == 486:
        return "Busy"
    if code == 404:
        return "Not found / unroutable"
    if code == 503:
        return "Service unavailable (after ring)" if rang else "Service unavailable - upstream (no ring)"
    if code == 500:
        return f"Server error: {reason or 'Internal Error'}"
    base = {401: "Unauthorized", 408: "Request timeout", 603: "Declined"}.get(code, "Failed")
    if reason:
        return f"Server error: {reason}"
    return f"{base} ({code})"


def _delta(a, b):
    """Whole seconds between two timestamps, or '' if either is missing."""
    if a and b:
        return round((b - a).total_seconds())
    return ""

def build_row(client, item):
    meta = client.get_json(f"/console/call-records/{item['id']}/metadata")
    start = _parse_ts(meta.get("startTime"))
    ring = _parse_ts(meta.get("ringTime"))
    answer = _parse_ts(meta.get("answerTime"))
    end = _parse_ts(meta.get("endTime"))
    rang = ring is not None

    # Timeline: start -> ring -> answer -> end
    total_secs = _delta(start, end)
    if answer:
        # ringing = answer - ring (fall back to start), talk = end - answer
        ring_secs = _delta(ring or start, answer)
        talk_secs = _delta(answer, end)
    else:
        # never answered: ring lasted until teardown, no talk time
        ring_secs = _delta(ring, end)
        talk_secs = 0
    return {
        "id": item["id"],
        # render the UTC start time in IST (UTC+5:30) for the human-facing CSV
        "started_at": start.astimezone(IST).strftime("%Y-%m-%d %H:%M:%S") if start else "",
        "direction": meta.get("direction", ""),
        "from": meta.get("fromNumber") or meta.get("caller") or "",
        "to": meta.get("toNumber") or meta.get("callee") or "",
        "status": meta.get("status", ""),
        "status_code": meta.get("statusCode", ""),
        "ring_secs": ring_secs,
        "talk_secs": talk_secs,
        "total_secs": total_secs,
        "error_reason": classify(meta, rang),
    }


def build_summary(rows, date_from, date_to):
    """Aggregate stats for the Summary sheet. Returns list of (label, value)."""
    total = len(rows)
    answered = [r for r in rows if r["status_code"] == 200 and r["status"] == "completed"]
    failed = [r for r in rows if r not in answered]
    inbound = sum(1 for r in rows if r["direction"] == "inbound")
    outbound = sum(1 for r in rows if r["direction"] == "outbound")
    talk = [r["talk_secs"] for r in answered if isinstance(r["talk_secs"], (int, float))]
    total_talk = sum(talk)
    asr = (len(answered) / total * 100) if total else 0.0

    rng = date_from if date_from == date_to else f"{date_from} .. {date_to}"
    summary = [
        ("Date range", rng or "all"),
        ("Total calls", total),
        ("Answered", len(answered)),
        ("Failed / unanswered", len(failed)),
        ("ASR (answer-seizure ratio)", f"{asr:.1f}%"),
        ("Inbound", inbound),
        ("Outbound", outbound),
        ("Total talk time (min)", round(total_talk / 60, 1)),
        ("Avg talk time, answered (s)", round(total_talk / len(talk), 1) if talk else 0),
        ("", ""),
        ("Outcome breakdown", "count"),
    ]
    breakdown = Counter(r["error_reason"] or "Answered" for r in rows)
    for reason, count in breakdown.most_common():
        summary.append((reason, count))
    return summary


PER_NUMBER_COLUMNS = ["Number", "Total", "Answered", "Rang no-answer", "Failed", "ASR", "Talk min"]


def build_per_number(rows):
    """One row of stats per caller number, sorted by call volume desc."""
    groups = {}
    for r in rows:
        groups.setdefault(r["from"] or "(unknown)", []).append(r)
    table = []
    for number, grp in groups.items():
        answered = [r for r in grp if r["status_code"] == 200 and r["status"] == "completed"]
        rang_noans = [r for r in grp if r["error_reason"] == "Rang but not answered"]
        talk = sum(r["talk_secs"] for r in answered if isinstance(r["talk_secs"], (int, float)))
        asr = (len(answered) / len(grp) * 100) if grp else 0.0
        table.append({
            "Number": number,
            "Total": len(grp),
            "Answered": len(answered),
            "Rang no-answer": len(rang_noans),
            "Failed": len(grp) - len(answered),
            "ASR": f"{asr:.1f}%",
            "Talk min": round(talk / 60, 1),
        })
    table.sort(key=lambda x: x["Total"], reverse=True)
    return table


def write_xlsx(path, rows, summary, per_number):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")

    # --- Summary sheet (first) ---
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["Metric", "Value"])
    for c in ws_sum[1]:
        c.font = header_font
        c.fill = header_fill
    for label, value in summary:
        ws_sum.append([label, value])
    # bold the section sub-header row ("Outcome breakdown")
    for row in ws_sum.iter_rows(min_row=2):
        if row[0].value == "Outcome breakdown":
            row[0].font = Font(bold=True)
            row[1].font = Font(bold=True)
    ws_sum.column_dimensions["A"].width = 34
    ws_sum.column_dimensions["B"].width = 18

    # --- Per-number breakdown table (below the metrics) ---
    ws_sum.append([])
    title_row = ws_sum.max_row + 1
    ws_sum.cell(row=title_row, column=1, value="Per-number breakdown").font = Font(bold=True)
    hdr_row = title_row + 1
    for col, name in enumerate(PER_NUMBER_COLUMNS, start=1):
        c = ws_sum.cell(row=hdr_row, column=col, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    for entry in per_number:
        ws_sum.append([entry[col] for col in PER_NUMBER_COLUMNS])

    # --- Calls sheet ---
    ws = wb.create_sheet("Calls")
    ws.append(CSV_COLUMNS)
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([r.get(col, "") for col in CSV_COLUMNS])
    widths = {"started_at": 20, "from": 16, "to": 16, "error_reason": 34}
    for idx, col in enumerate(CSV_COLUMNS, start=1):
        ws.column_dimensions[chr(64 + idx)].width = widths.get(col, 12)
    ws.freeze_panes = "A2"

    wb.save(path)


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("-n", "--number", action="append", default=[],
                        help="phone number to search for (repeatable). Empty = all calls.")
    parser.add_argument("--date", help="single day (YYYY-MM-DD); sets both from and to")
    parser.add_argument("--date-from", help="start date YYYY-MM-DD")
    parser.add_argument("--date-to", help="end date YYYY-MM-DD")
    parser.add_argument("-o", "--out", default="call_records.csv",
                        help="output path. .xlsx => two-sheet workbook (Summary + Calls); "
                             ".csv => records CSV plus a *_summary.csv companion")
    parser.add_argument("--base-url", default=DEFAULT_BASE)
    parser.add_argument("--identifier", default=DEFAULT_USER)
    parser.add_argument("--password", default=DEFAULT_PASS)
    args = parser.parse_args(argv)

    date_from = args.date or args.date_from
    date_to = args.date or args.date_to

    client = Client(args.base_url, args.identifier, args.password).login()

    # de-dupe by record id across the multiple number searches
    by_id = {}
    search_terms = args.number or [""]
    for term in search_terms:
        for item in fetch_records(client, term, date_from, date_to):
            by_id[item["id"]] = item
    print(f"found {len(by_id)} records; fetching metadata...", file=sys.stderr)

    rows = [build_row(client, item) for item in by_id.values()]
    rows.sort(key=lambda r: r["started_at"], reverse=True)

    summary = build_summary(rows, date_from, date_to)
    per_number = build_per_number(rows)

    if args.out.lower().endswith(".xlsx"):
        write_xlsx(args.out, rows, summary, per_number)
    else:
        with open(args.out, "w", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=CSV_COLUMNS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
        # also drop a companion summary CSV next to it
        sum_path = re.sub(r"\.csv$", "", args.out, flags=re.I) + "_summary.csv"
        with open(sum_path, "w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["Metric", "Value"])
            w.writerows(summary)
            w.writerow([])
            w.writerow(["Per-number breakdown"])
            w.writerow(PER_NUMBER_COLUMNS)
            for entry in per_number:
                w.writerow([entry[col] for col in PER_NUMBER_COLUMNS])
        print(f"wrote {sum_path}", file=sys.stderr)

    print(f"wrote {args.out} ({len(rows)} rows)", file=sys.stderr)
    for label, value in summary:
        if label and label not in ("Outcome breakdown",):
            print(f"  {label}: {value}", file=sys.stderr)


if __name__ == "__main__":
    main()
